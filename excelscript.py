import firebase_admin
from firebase_admin import credentials
from firebase_admin import db
from openpyxl import load_workbook

cred = credentials.Certificate("./bakeoff2-5004a-firebase-adminsdk-mx2d0-0d834c410e.json")
firebase_admin.initialize_app(cred, {'databaseURL': 'https://bakeoff2-5004a-default-rtdb.europe-west1.firebasedatabase.app/'})

ref = db.reference("/G47")

# Retrieve data from the Firebase database
data = ref.get()

# Create workbook and worksheet
wb = load_workbook(filename="Resultados.xlsx")
ws = wb.active

key_list = []

for cell in ws["A"]:
    key_list.append(cell.value)

# Loop through data and add rows to worksheet
for key, value in data.items():
    if key not in key_list:
        ws.append([key,  1, value["accuracy"], str(value["assessed_by"]),
        value["attempt"], value["attempt_duration"], value["hits"],
        value["misses"], float(value["target_w_penalty"]), value["test_completed_by"],
        float(value["time_per_target"])])


# Save workbook to file
wb.save("Resultados.xlsx")

# Make sure to close the workbook
wb.close()